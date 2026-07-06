"""
Build a live Excel cost-benefit workbook. All damages/exposure come in as values
(computed from the GIS pipeline); AED, NPV and BCR are Excel FORMULAS that reference
the Inputs sheet, so editing project cost / discount rate / maintenance updates
everything automatically.
Output: output/Majlas_CBA.xlsx
"""
import os, csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import config as C

BOLD = Font(bold=True)
HFILL = PatternFill("solid", fgColor="333333"); HFONT = Font(bold=True, color="FFFFFF")
INFILL = PatternFill("solid", fgColor="FFF2CC")   # editable inputs
OUTFILL = PatternFill("solid", fgColor="DDEBF7")
thin = Side(style="thin", color="BBBBBB"); BORD = Border(thin, thin, thin, thin)
RPS = C.RP_YEARS


def _read():
    r = list(csv.DictReader(open(os.path.join(C.OUT_DIR, "damage_by_plan.csv"))))
    b = {int(x["rp"]): float(x["total"]) for x in r if x["condition"] == "baseline"}
    s = {int(x["rp"]): float(x["total"]) for x in r if x["condition"] == "scheme"}
    e = {(int(x["rp"]), x["condition"]): x
         for x in csv.DictReader(open(os.path.join(C.OUT_DIR, "exposure_by_rp.csv")))}
    return b, s, e


def hdr(ws, row, cols):
    for j, t in enumerate(cols, 1):
        c = ws.cell(row, j, t); c.fill = HFILL; c.font = HFONT
        c.alignment = Alignment(horizontal="center"); c.border = BORD


def main():
    Db, Ds, EX = _read()
    wb = openpyxl.Workbook()

    # ---------------- Inputs ----------------
    ip = wb.active; ip.title = "Inputs"
    ip["A1"] = "MAJLAS FLOOD PROTECTION — COST-BENEFIT INPUTS"; ip["A1"].font = Font(bold=True, size=13)
    rows = [("Dam capital cost", 75_000_000, "OMR"),
            ("Channel capital cost", 5_000_000, "OMR"),
            ("Relocation cost", 168_080, "OMR"),
            ("Discount rate (r)", 0.025, "fraction"),
            ("Horizon (n)", 50, "years"),
            ("Dam maintenance rate", 0.01, "of dam capital / yr"),
            ("Dam maintenance start", 5, "year"),
            ("Channel maintenance rate", 0.007, "of channel capital / yr"),
            ("Channel maintenance start", 2, "year")]
    ip.append([]); hdr(ip, 3, ["Parameter", "Value", "Unit"])
    for i, (lbl, val, unit) in enumerate(rows, 4):
        ip.cell(i, 1, lbl); v = ip.cell(i, 2, val); ip.cell(i, 3, unit)
        v.fill = INFILL; v.border = BORD
        if "rate" in lbl.lower(): v.number_format = "0.0%"
        elif "cost" in lbl.lower(): v.number_format = "#,##0"
    ip.cell(14, 1, "↑ yellow cells are editable — results update automatically").font = Font(italic=True)
    DAM, CHAN, RELOC, R, N = "Inputs!$B$4", "Inputs!$B$5", "Inputs!$B$6", "Inputs!$B$7", "Inputs!$B$8"
    MDR, MDS, MCR, MCS = "Inputs!$B$9", "Inputs!$B$10", "Inputs!$B$11", "Inputs!$B$12"
    ip.column_dimensions["A"].width = 26; ip.column_dimensions["C"].width = 22

    # ---------------- Damage & AED ----------------
    da = wb.create_sheet("Damage_AED")
    da["A1"] = "Flood damage by return period  (from GIS pipeline)"; da["A1"].font = BOLD
    hdr(da, 3, ["Return period (yr)", "Exceedance prob", "Damage baseline (OMR)",
                "Damage scheme (OMR)", "Trapezoid baseline", "Trapezoid scheme"])
    r0 = 4
    for i, rp in enumerate(RPS):
        r = r0 + i
        da.cell(r, 1, rp)
        da.cell(r, 2, f"=1/A{r}").number_format = "0.0000"
        da.cell(r, 3, round(Db[rp])).number_format = "#,##0"
        da.cell(r, 4, round(Ds[rp])).number_format = "#,##0"
        if i < len(RPS) - 1:
            da.cell(r, 5, f"=0.5*(C{r}+C{r+1})*(B{r}-B{r+1})").number_format = "#,##0"
            da.cell(r, 6, f"=0.5*(D{r}+D{r+1})*(B{r}-B{r+1})").number_format = "#,##0"
        for j in range(1, 7): da.cell(r, j).border = BORD
    last = r0 + len(RPS) - 1
    # AED = trapezoid sum + flat tail (rarest damage * smallest prob); zero below 2yr
    base_r, sch_r, av_r = last + 2, last + 3, last + 4
    da.cell(base_r, 1, "AED baseline (OMR/yr)").font = BOLD
    da.cell(base_r, 2, f"=SUM(E{r0}:E{last-1})+C{last}*B{last}").number_format = "#,##0"
    da.cell(sch_r, 1, "AED scheme (OMR/yr)").font = BOLD
    da.cell(sch_r, 2, f"=SUM(F{r0}:F{last-1})+D{last}*B{last}").number_format = "#,##0"
    da.cell(av_r, 1, "Annual avoided damage (OMR/yr)").font = BOLD
    av_cell = f"Damage_AED!$B${av_r}"
    da.cell(av_r, 2, f"=B{base_r}-B{sch_r}").number_format = "#,##0"
    da.cell(av_r, 2).fill = OUTFILL
    da.column_dimensions["A"].width = 30
    for col in "BCDEF": da.column_dimensions[col].width = 18

    # ---------------- Cashflow / NPV / BCR ----------------
    cf = wb.create_sheet("NPV_BCR")
    cf["A1"] = "Discounted cash flow (formula-driven)"; cf["A1"].font = BOLD
    hdr(cf, 3, ["Year", "Avoided damage", "Dam maint.", "Channel maint.",
                "Net benefit", "Discount factor", "PV net", "PV benefit", "PV cost"])
    y0 = 4
    for t in range(1, 51):
        r = y0 + t - 1
        cf.cell(r, 1, t)
        cf.cell(r, 2, f"={av_cell}")
        cf.cell(r, 3, f"=IF(A{r}>={MDS},{MDR}*{DAM},0)")
        cf.cell(r, 4, f"=IF(A{r}>={MCS},{MCR}*{CHAN},0)")
        cf.cell(r, 5, f"=B{r}-C{r}-D{r}")
        cf.cell(r, 6, f"=1/(1+{R})^A{r}").number_format = "0.000"
        cf.cell(r, 7, f"=E{r}*F{r}")
        cf.cell(r, 8, f"=B{r}*F{r}")
        cf.cell(r, 9, f"=(C{r}+D{r})*F{r}")
        for j in range(2, 10):
            if j != 6: cf.cell(r, j).number_format = "#,##0"
    ylast = y0 + 49
    lbls = [("Initial cost (dam+channel+relocation)", f"={DAM}+{CHAN}+{RELOC}"),
            ("PV of benefits (avoided damage)", f"=SUM(H{y0}:H{ylast})"),
            ("PV of maintenance", f"=SUM(I{y0}:I{ylast})"),
            ("PV of costs (initial + maintenance)", None),
            ("NPV", None), ("BCR", None)]
    base = ylast + 2
    for k, (lbl, f) in enumerate(lbls):
        r = base + k; cf.cell(r, 1, lbl).font = BOLD
        cell = cf.cell(r, 2)
        if f: cell.value = f
        cell.number_format = "#,##0"
    IC, PVB, PVM = f"B{base}", f"B{base+1}", f"B{base+2}"
    cf.cell(base+3, 2, f"={IC}+{PVM}").number_format = "#,##0"          # PV costs
    cf.cell(base+4, 2, f"={PVB}-{PVM}-{IC}").number_format = "#,##0"    # NPV
    cf.cell(base+4, 2).fill = OUTFILL; cf.cell(base+4, 2).font = BOLD
    cf.cell(base+5, 2, f"={PVB}/B{base+3}").number_format = "0.000"      # BCR
    cf.cell(base+5, 2).fill = OUTFILL; cf.cell(base+5, 2).font = BOLD
    cf.column_dimensions["A"].width = 34
    for col in "BCDEFGHI": cf.column_dimensions[col].width = 15

    # ---------------- Exposure ----------------
    ex = wb.create_sheet("Exposure")
    ex["A1"] = "Population exposure by return period"; ex["A1"].font = BOLD
    hdr(ex, 3, ["Return period (yr)", "Prob", "People exposed base", "People exposed scheme",
                ">1 m base (life-risk)", ">1 m scheme", "Trap base", "Trap scheme"])
    for i, rp in enumerate(RPS):
        r = 4 + i
        ex.cell(r, 1, rp); ex.cell(r, 2, f"=1/A{r}").number_format = "0.0000"
        ex.cell(r, 3, int(float(EX[(rp,'baseline')]["people_exposed"])))
        ex.cell(r, 4, int(float(EX[(rp,'scheme')]["people_exposed"])))
        ex.cell(r, 5, int(float(EX[(rp,'baseline')]["people_gt1.0m"])))
        ex.cell(r, 6, int(float(EX[(rp,'scheme')]["people_gt1.0m"])))
        if i < len(RPS)-1:
            ex.cell(r, 7, f"=0.5*(C{r}+C{r+1})*(B{r}-B{r+1})").number_format = "#,##0"
            ex.cell(r, 8, f"=0.5*(D{r}+D{r+1})*(B{r}-B{r+1})").number_format = "#,##0"
    el = 4 + len(RPS) - 1
    ex.cell(el+2, 1, "EAPE baseline (people/yr)").font = BOLD
    ex.cell(el+2, 2, f"=SUM(G4:G{el-1})+C{el}*B{el}").number_format = "#,##0"
    ex.cell(el+3, 1, "EAPE scheme (people/yr)").font = BOLD
    ex.cell(el+3, 2, f"=SUM(H4:H{el-1})+D{el}*B{el}").number_format = "#,##0"
    ex.cell(el+4, 1, "Annual people-exposure avoided").font = BOLD
    ex.cell(el+4, 2, f"=B{el+2}-B{el+3}").number_format = "#,##0"; ex.cell(el+4,2).fill = OUTFILL
    ex.column_dimensions["A"].width = 26
    for col in "BCDEFGH": ex.column_dimensions[col].width = 16

    # ---------------- Summary ----------------
    sm = wb.create_sheet("Summary"); wb.move_sheet("Summary", -(len(wb.sheetnames)-1))
    sm["A1"] = "MAJLAS FLOOD PROTECTION — COST-BENEFIT SUMMARY"; sm["A1"].font = Font(bold=True, size=14)
    items = [("Annual avoided damage (OMR/yr)", f"={av_cell}", "#,##0"),
             ("AED baseline (OMR/yr)", f"=Damage_AED!$B${base_r}", "#,##0"),
             ("AED scheme (OMR/yr)", f"=Damage_AED!$B${sch_r}", "#,##0"),
             ("Initial cost (OMR)", f"=NPV_BCR!{IC}", "#,##0"),
             ("PV benefits (OMR)", f"=NPV_BCR!{PVB}", "#,##0"),
             ("PV costs (OMR)", f"=NPV_BCR!B{base+3}", "#,##0"),
             ("NPV (OMR)", f"=NPV_BCR!B{base+4}", "#,##0"),
             ("BCR", f"=NPV_BCR!B{base+5}", "0.000"),
             ("EAPE avoided (people/yr)", f"=Exposure!$B${el+4}", "#,##0")]
    hdr(sm, 3, ["Metric", "Value"])
    for i, (lbl, f, fmt) in enumerate(items, 4):
        sm.cell(i, 1, lbl); c = sm.cell(i, 2, f); c.number_format = fmt; c.border = BORD
        sm.cell(i, 1).border = BORD
        if lbl in ("NPV (OMR)", "BCR"): c.fill = OUTFILL; c.font = BOLD; sm.cell(i,1).font = BOLD
    sm.cell(14, 1, "Verdict:").font = BOLD
    sm.cell(14, 2, '=IF(B11>1,"Viable (BCR>1)","Not viable on damage alone - see life-safety metrics")')
    sm.column_dimensions["A"].width = 34; sm.column_dimensions["B"].width = 20

    out = os.path.join(C.OUT_DIR, "Majlas_CBA.xlsx")
    wb.save(out); print("wrote", out)


if __name__ == "__main__":
    main()
